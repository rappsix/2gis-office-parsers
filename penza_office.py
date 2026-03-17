import ssl
import time
import re
import os
import random
import requests
import urllib.parse
from urllib.parse import quote, urljoin, urlparse

# === ПАТЧ ДЛЯ MAC OS (SSL) ===
try:
    _create_unverified_https_context = ssl._create_unverified_context
except AttributeError:
    pass
else:
    ssl._create_default_https_context = _create_unverified_https_context

from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
import undetected_chromedriver as uc
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains

# ================= НАСТРОЙКИ =================
CITY_URL = "https://2gis.ru/penza"
OUTPUT_FILE = "2gis_penza_office_service_dedup.xlsx"
MAX_PAGES_PER_CATEGORY = 32
WEB_TIMEOUT = 15

# ================= КАТЕГОРИИ ПОД ОФИСНУЮ ТЕМУ =================
CATEGORIES = [
    "ремонт офисной техники",
    "ремонт оргтехники",
    "ремонт принтеров",
    "ремонт МФУ",
    "ремонт копиров",
    "обслуживание офисной техники",
    "обслуживание оргтехники",
    "сервисный центр оргтехники",
    "сервисный центр офисной техники",

    "ремонт кофемашин",
    "обслуживание кофемашин",
    "сервисный центр кофемашин",
    "установка кофемашин",
    "аренда кофемашин",
    "кофемашины для офиса",

    "копировальные центры",
    "копировальный центр",
    "центр копирования и печати",
    "типография цифровая печать",
    "услуги печати документов",
    "печать и ксерокопия",
    "ксерокопия и распечатка",

    "обслуживание компьютеров для организаций",
    "аутсорсинг ИТ",
    "ИТ обслуживание организаций",
    "обслуживание компьютеров и оргтехники",
    "настройка и ремонт компьютеров для бизнеса",

    "заправка картриджей",
    "ремонт картриджей",
    "продажа офисной техники",
    "магазин офисной техники",
    "аренда принтеров",
    "аренда копиров",

    "офисное оборудование",
    "поставка офисной техники",
    "обслуживание предприятий оргтехникой",
    "услуги для офиса",
]
CATEGORIES = list(dict.fromkeys(CATEGORIES))

IGNORE_EMAILS = [
    "support@maps.yandex.ru", "abuse@yandex.ru", "noreply@yandex.ru",
    "no-reply@yandex.ru", "info@yandex.ru", "mail@yandex.ru",
    "info@2gis.ru", "support@2gis.ru", "example@example.com",
    "sentry@", "wixpress.com"
]
EMAIL_RE = re.compile(r"[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}")

def try_auto_solve_captcha(driver):
    print("   🤖 [AUTO] Пробую пройти капчу...")
    try:
        iframes = driver.find_elements(By.TAG_NAME, "iframe")
        for frame in iframes:
            try:
                if frame.size['width'] > 0:
                    driver.switch_to.frame(frame)
                    checkbox = driver.find_elements(
                        By.CSS_SELECTOR,
                        "input[type='checkbox'], div.recaptcha-checkbox-border"
                    )
                    if checkbox:
                        el = checkbox[0]
                        action = ActionChains(driver)
                        action.move_to_element_with_offset(
                            el,
                            random.randint(2, 5),
                            random.randint(2, 5)
                        )
                        action.pause(0.3)
                        action.click()
                        action.perform()
                        time.sleep(2)
                        driver.switch_to.default_content()
                        return True
                    driver.switch_to.default_content()
            except:
                driver.switch_to.default_content()

        buttons = driver.find_elements(
            By.XPATH,
            "//button[contains(text(), 'человек') or contains(text(), 'Verify')]"
        )
        if buttons:
            buttons[0].click()
            time.sleep(2)
            return True
    except:
        pass
    return False

def check_captcha_smart(driver):
    if "captcha" in driver.current_url.lower() or "challenge" in driver.current_url.lower():
        print("\n🛑 КАПЧА!")
        try_auto_solve_captcha(driver)
        time.sleep(3)
        if "captcha" not in driver.current_url.lower():
            print("✅ Пройдено.")
            return
        print("⚠️ Реши вручную и нажми Enter.")
        input()

def search_site_duckduckgo(company_name, city="Пенза"):
    if not company_name:
        return None
    try:
        query = f"{company_name} {city} официальный сайт"
        url = f"https://html.duckduckgo.com/html/?q={quote(query)}"
        headers = {"User-Agent": "Mozilla/5.0"}
        resp = requests.get(url, headers=headers, timeout=10)
        soup = BeautifulSoup(resp.text, 'html.parser')

        for link in soup.select(".result__a")[:3]:
            href = link.get('href')
            if not href:
                continue

            if "duckduckgo.com/l/?" in href:
                try:
                    parsed = urllib.parse.parse_qs(urllib.parse.urlparse(href).query)
                    real_url = parsed.get('uddg', [None])[0]
                except:
                    real_url = href
            else:
                real_url = href

            if not real_url:
                continue

            skip = [
                'yandex', '2gis', 'zoon', 'vk.com', 'avito', 'hh.ru',
                'instagram.com', 'ok.ru', 't.me'
            ]
            if not any(d in real_url for d in skip):
                return real_url
    except:
        pass
    return None

def extract_emails_from_html(text):
    emails = set(EMAIL_RE.findall(text))
    valid = []
    for em in emails:
        el = em.lower()
        if not any(x in el for x in ['.png', '.jpg', '.js', 'sentry']) and el not in IGNORE_EMAILS:
            valid.append(em)
    if not valid:
        return ""
    valid.sort(key=lambda x: 0 if any(p in x.lower() for p in ['info', 'mail']) else 1)
    return valid[0]

def find_emails_on_site_recursive(url):
    if not url:
        return ""
    headers = {"User-Agent": "Mozilla/5.0"}
    to_visit = [url]

    # 1. Главная страница
    try:
        resp = requests.get(url, headers=headers, timeout=WEB_TIMEOUT, verify=False)
        resp.encoding = resp.apparent_encoding
        text = resp.text
        first_email = extract_emails_from_html(text)
        if first_email:
            return first_email

        soup = BeautifulSoup(text, 'html.parser')
        # Ищем ссылки на контакты / about и т.п.
        candidate_links = []
        for a in soup.find_all('a', href=True):
            href = a['href']
            anchor_text = a.get_text().lower()
            if any(k in anchor_text for k in ['контакт', 'связаться', 'contact']) or \
               any(k in href.lower() for k in ['contact', 'about', 'kontakty', 'kontact']):
                full = urljoin(url, href)
                if urlparse(full).netloc == urlparse(url).netloc and full not in candidate_links:
                    candidate_links.append(full)
        # ограничиваемся максимум 2‑мя такими страницами
        for link in candidate_links[:2]:
            to_visit.append(link)
    except:
        pass

    # 2. Дополнительные страницы (контакты)
    for link in to_visit[1:]:
        try:
            c_resp = requests.get(link, headers=headers, timeout=WEB_TIMEOUT, verify=False)
            c_resp.encoding = c_resp.apparent_encoding
            email = extract_emails_from_html(c_resp.text)
            if email:
                return email
        except:
            continue

    return ""

def search_email_yandex_selenium(driver, company_name):
    try:
        query = company_name + " Пенза email почта"
        driver.get(f"https://yandex.ru/search/?text={quote(query)}")
        time.sleep(1.5)
        check_captcha_smart(driver)
        try:
            body = driver.find_element(By.TAG_NAME, "body").text
        except:
            return ""
        return extract_emails_from_html(body)
    except:
        pass
    return ""

def setup_driver():
    options = uc.ChromeOptions()
    # options.add_argument("--headless=new")  # убрали headless, видимый браузер
    options.add_argument("--start-maximized")
    options.add_argument("--disable-web-security")
    options.add_argument("--disable-features=VizDisplayCompositor")
    print("🚀 Запуск браузера (видимый режим)...")
    driver = uc.Chrome(options=options)
    return driver


def main():
    requests.packages.urllib3.disable_warnings()
    driver = setup_driver()

    # Excel
    if not os.path.exists(OUTPUT_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(["Категория", "Название", "Телефон", "Сайт", "Email", "Источник", "Ссылка"])
        for cell in ws[1]:
            cell.fill = PatternFill(start_color="006400", end_color="006400", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        wb.save(OUTPUT_FILE)
    else:
        wb = load_workbook(OUTPUT_FILE)
        ws = wb.active

    processed_urls = set()
    seen_keys = set()  # фильтр по дублям

    print(f"✅ Старт. Категорий: {len(CATEGORIES)}")

    for idx, category in enumerate(CATEGORIES, 1):
        print(f"\n[{idx}/{len(CATEGORIES)}] 📂 {category}")

        try:
            driver.get(f"{CITY_URL}/search/{quote(category)}")
            time.sleep(4)
            check_captcha_smart(driver)
        except:
            continue

        for page_num in range(1, MAX_PAGES_PER_CATEGORY + 1):
            print(f"   📄 Стр. {page_num}...", end=" ")

            links_on_page = []
            for attempt in range(3):
                try:
                    elements = driver.find_elements(
                        By.CSS_SELECTOR,
                        "a[href^='/penza/firm/'], a[href^='/penza/geo/']"
                    )
                    found = []
                    for el in elements:
                        href = el.get_attribute("href")
                        if not href:
                            continue
                        href = href.split("?")[0]
                        if "/penza/firm/" in href or "/penza/geo/" in href:
                            found.append(href)

                    new_links = [L for L in found if L not in processed_urls]
                    if new_links:
                        links_on_page = new_links
                        for L in new_links:
                            processed_urls.add(L)
                        break
                    time.sleep(2)
                except:
                    time.sleep(1)

            if not links_on_page:
                print("Пусто.")
                if page_num > 1:
                    print("   ⏹️ Похоже, конец списка.")
                    break
                else:
                    break

            print(f"Новых: {len(links_on_page)}")

            for link in links_on_page:
                try:
                    driver.get(link)
                    time.sleep(0.8)
                    # Название
                    try:
                        name = driver.find_element(By.TAG_NAME, "h1").text.strip()
                    except:
                        continue
                    if not name:
                        continue

                    # Телефон
                    phone = ""
                    try:
                        phones = driver.find_elements(By.CSS_SELECTOR, "a[href^='tel:']")
                        phone = ", ".join(sorted(set(
                            [p.text.strip() for p in phones if p.text.strip()]
                        )))
                    except:
                        pass

                    # Сайт
                    site = ""
                    try:
                        sites = driver.find_elements(By.CSS_SELECTOR, "a[target='_blank']")
                        for s in sites:
                            h = s.get_attribute("href")
                            if not h:
                                continue
                            lower = h.lower()
                            if any(bad in lower for bad in [
                                "2gis.ru", "vk.com", "ok.ru", "instagram.com", "t.me"
                            ]):
                                continue
                            site = h
                            break
                    except:
                        pass

                    email = ""
                    source = ""

                    # 1. Email из 2ГИС
                    try:
                        mails = driver.find_elements(By.CSS_SELECTOR, "a[href^='mailto:']")
                        if mails:
                            email = mails[0].get_attribute("href").replace("mailto:", "").strip()
                            source = "2ГИС"
                    except:
                        pass

                    # 2. Скан сайта (главная + контакты)
                    if not email:
                        if not site:
                            site = search_site_duckduckgo(name, "Пенза")
                        if site:
                            email = find_emails_on_site_recursive(site)
                            if email:
                                source = "Сайт (скан)"

                    # 3. Яндекс
                    if not email:
                        email = search_email_yandex_selenium(driver, name)
                        if email:
                            source = "Яндекс"

                    # --- ФИЛЬТР ДУБЛЕЙ ---
                    if email:
                        key = ("email", email.lower())
                    else:
                        key = ("name_phone", name.lower(), phone)

                    if key in seen_keys:
                        print(f"     SKIP (дубль) {name}")
                        continue
                    seen_keys.add(key)

                    print(f"     {name} | 📞 {phone or '-'} | 📧 {email or '-'} | 🌐 {site or '-'}")
                    ws.append([category, name, phone, site, email, source, link])
                    wb.save(OUTPUT_FILE)

                except:
                    pass

            # Переход на следующую страницу
            try:
                next_url = f"{CITY_URL}/search/{quote(category)}/page/{page_num + 1}"
                driver.get(next_url)
                time.sleep(3)
                check_captcha_smart(driver)
                if f"/page/{page_num + 1}" not in driver.current_url:
                    print(f"   ⏹️ Редирект (конец).")
                    break
            except:
                break

    driver.quit()
    print(f"✅ Готово! Файл: {OUTPUT_FILE}")

if __name__ == "__main__":
    main()
