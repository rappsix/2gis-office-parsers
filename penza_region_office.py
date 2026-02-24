import ssl
import time
import re
import os
import random
import requests
import urllib.parse
from urllib.parse import quote, urljoin, urlparse
import undetected_chromedriver as uc
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys

# === ПАТЧ SSL ===
try:
    _create_unverified_https_context = ssl._create_unverified_context
except AttributeError:
    pass
else:
    ssl._create_default_https_context = _create_unverified_https_context

# ================= НАСТРОЙКИ =================
BASE_URL = "https://2gis.ru/penza"
OUTPUT_FILE = "2gis_penza_region_office_dedup.xlsx"
MAX_PAGES_PER_QUERY = 5
WEB_TIMEOUT = 10
REGION_NAME = "Пензенская область"

# ================= ГОРОДА / РАЙОНЫ =================
AREA_SYSTEM = [
    # сюда можешь вернуть системные города
    # "Кузнецк", "Заречный", ...
]
AREA_CUSTOM = [
    "Кижеватово", "Вазерки",
]
TARGET_AREAS = list(dict.fromkeys(AREA_SYSTEM + AREA_CUSTOM))

# ================= КАТЕГОРИИ (офисная тематика) =================
CATEGORIES = [
    "ремонт офисной техники",
    "ремонт оргтехники",
    "ремонт принтеров",
    "ремонт МФУ",
    "ремонт копиров",
    "обслуживание офисной техники",
    "обслуживание оргтехники",
    "сервисный центр оргтехники",
    "сервисный центр офисной техники",

    "ремонт кофемашин",
    "обслуживание кофемашин",
    "сервисный центр кофемашин",
    "установка кофемашин",
    "аренда кофемашин",
    "кофемашины для офиса",

    "копировальные центры",
    "копировальный центр",
    "центр копирования и печати",
    "типография цифровая печать",
    "услуги печати документов",
    "печать и ксерокопия",
    "ксерокопия и распечатка",

    "обслуживание компьютеров для организаций",
    "аутсорсинг ИТ",
    "ИТ обслуживание организаций",
    "обслуживание компьютеров и оргтехники",
    "настройка и ремонт компьютеров для бизнеса",

    "заправка картриджей",
    "ремонт картриджей",
    "продажа офисной техники",
    "магазин офисной техники",
    "аренда принтеров",
    "аренда копиров",

    "офисное оборудование",
    "поставка офисной техники",
    "обслуживание предприятий оргтехникой",
    "услуги для офиса",
]
CATEGORIES = list(dict.fromkeys(CATEGORIES))

IGNORE_EMAILS = [
    "support@maps.yandex.ru", "abuse@yandex.ru", "noreply@yandex.ru",
    "no-reply@yandex.ru", "info@yandex.ru", "mail@yandex.ru",
    "info@2gis.ru", "support@2gis.ru", "example@example.com", "sentry@", "wixpress.com"
]
EMAIL_RE = re.compile(r"[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}")

def try_auto_solve_captcha(driver):
    try:
        iframes = driver.find_elements(By.TAG_NAME, "iframe")
        for frame in iframes:
            try:
                if frame.size['width'] > 0:
                    driver.switch_to.frame(frame)
                    checkbox = driver.find_elements(By.CSS_SELECTOR, "input[type='checkbox'], div.recaptcha-checkbox-border")
                    if checkbox:
                        checkbox[0].click()
                        driver.switch_to.default_content()
                        return True
                    driver.switch_to.default_content()
            except:
                driver.switch_to.default_content()
        buttons = driver.find_elements(By.XPATH, "//button[contains(text(), 'человек') or contains(text(), 'Verify') or contains(text(), 'робот') or contains(text(), 'нажмите') ]")
        if buttons:
            buttons[0].click()
            time.sleep(2)
            return True
    except:
        pass
    return False

def check_captcha_smart(driver):
    if "captcha" in driver.current_url.lower() or "challenge" in driver.current_url.lower():
        print("\n🛑 КАПЧА! (Пауза 5с)")
        time.sleep(5)
        try_auto_solve_captcha(driver)
        time.sleep(3)
        if "captcha" not in driver.current_url.lower():
            return
        print("⚠️ Реши вручную и нажми Enter.")
        input()

def search_site_duckduckgo(company_name, city):
    if not company_name:
        return None
    try:
        query = f"{company_name} {city} официальный сайт"
        url = f"https://html.duckduckgo.com/html/?q={quote(query)}"
        headers = {"User-Agent": "Mozilla/5.0"}
        resp = requests.get(url, headers=headers, timeout=10)
        soup = BeautifulSoup(resp.text, 'html.parser')
        for link in soup.select(".result__a")[:3]:
            href = link.get('href')
            if href:
                if "duckduckgo.com/l/?" in href:
                    try:
                        parsed = urllib.parse.parse_qs(urllib.parse.urlparse(href).query)
                        real_url = parsed.get('uddg', [None])[0]
                    except:
                        real_url = href
                else:
                    real_url = href
                skip = ['yandex', '2gis', 'zoon', 'vk.com', 'avito', 'hh.ru', 'instagram.com', 'ok.ru', 't.me']
                if real_url and not any(d in real_url for d in skip):
                    return real_url
    except:
        pass
    return None

def extract_emails_from_html(text):
    emails = set(EMAIL_RE.findall(text))
    valid = []
    for em in emails:
        el = em.lower()
        if not any(x in el for x in ['.png', '.jpg', '.js', 'sentry']) and el not in IGNORE_EMAILS:
            valid.append(em)
    if not valid:
        return ""
    valid.sort(key=lambda x: 0 if any(p in x.lower() for p in ['info', 'mail']) else 1)
    return valid[0]

def find_emails_on_site_recursive(url):
    if not url:
        return ""
    headers = {"User-Agent": "Mozilla/5.0"}
    to_visit = [url]

    try:
        resp = requests.get(url, headers=headers, timeout=WEB_TIMEOUT, verify=False)
        resp.encoding = resp.apparent_encoding
        text = resp.text
        first_email = extract_emails_from_html(text)
        if first_email:
            return first_email

        soup = BeautifulSoup(text, 'html.parser')
        candidate_links = []
        for a in soup.find_all('a', href=True):
            href = a['href']
            anchor_text = a.get_text().lower()
            if any(k in anchor_text for k in ['контакт', 'связаться', 'contact']) or \
               any(k in href.lower() for k in ['contact', 'about', 'kontakty', 'kontact']):
                full = urljoin(url, href)
                if urlparse(full).netloc == urlparse(url).netloc and full not in candidate_links:
                    candidate_links.append(full)
        for link in candidate_links[:2]:
            to_visit.append(link)
    except:
        pass

    for link in to_visit[1:]:
        try:
            c_resp = requests.get(link, headers=headers, timeout=WEB_TIMEOUT, verify=False)
            c_resp.encoding = c_resp.apparent_encoding
            email = extract_emails_from_html(c_resp.text)
            if email:
                return email
        except:
            continue

    return ""

def search_email_yandex_selenium(driver, query):
    try:
        driver.get(f"https://yandex.ru/search/?text={quote(query + ' email')}")
        time.sleep(1.5)
        check_captcha_smart(driver)
        try:
            body = driver.find_element(By.TAG_NAME, "body").text
        except:
            return ""
        return extract_emails_from_html(body)
    except:
        pass
    return ""

def setup_driver():
    options = uc.ChromeOptions()
    options.add_argument("--headless=new")  # без окна
    options.add_argument("--no-sandbox")    # для root и серверов
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-gpu")
    options.add_argument("--disable-extensions")
    options.add_argument("--start-maximized")
    options.add_argument("--disable-background-timer-throttling")
    options.add_argument("--disable-backgrounding-occluded-windows")
    options.add_argument("--disable-renderer-backgrounding")
    print("🚀 Запуск браузера (headless)...")
    driver = uc.Chrome(options=options)
    return driver

def main():
    requests.packages.urllib3.disable_warnings()
    driver = setup_driver()

    if not os.path.exists(OUTPUT_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(["Город", "Категория", "Название", "Телефон", "Сайт", "Email", "Источник", "Ссылка"])
        for cell in ws[1]:
            cell.fill = PatternFill(start_color="006400", end_color="006400", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        wb.save(OUTPUT_FILE)
        processed_urls = set()
    else:
        wb = load_workbook(OUTPUT_FILE)
        ws = wb.active
        processed_urls = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[7]:
                processed_urls.add(row[7])

    seen_keys = set()

    print(f"✅ ЗАПУСК ПО ОБЛАСТИ. Городов/районов: {len(TARGET_AREAS)}")

    for area in TARGET_AREAS:
        print(f"\n🌍 РАЙОН: {area}")

        for category in CATEGORIES:
            search_query = f"{category} {area} {REGION_NAME}"

            time.sleep(random.uniform(4.0, 7.0))
            try:
                driver.get(f"{BASE_URL}/search/{quote(search_query)}")
                time.sleep(2)
                ActionChains(driver).send_keys(Keys.PAGE_DOWN).perform()
                check_captcha_smart(driver)
            except:
                continue

            for page_num in range(1, MAX_PAGES_PER_QUERY + 1):
                links_on_page = []
                try:
                    elements = driver.find_elements(By.CSS_SELECTOR, "a[href*='/firm/'], a[href*='/geo/']")
                    for el in elements:
                        h = el.get_attribute("href")
                        if h:
                            clean_h = h.split('?')[0]
                            if clean_h not in processed_urls:
                                links_on_page.append(clean_h)
                                processed_urls.add(clean_h)
                    links_on_page = list(set(links_on_page))
                except:
                    pass

                if not links_on_page:
                    break

                print(f"   [{category}] +{len(links_on_page)} шт.")

                for link in links_on_page:
                    try:
                        driver.get(link)
                        time.sleep(0.8)
                        check_captcha_smart(driver)

                        try:
                            name = driver.find_element(By.TAG_NAME, "h1").text.strip()
                        except:
                            continue
                        if not name:
                            continue

                        phone = ""
                        try:
                            phones = driver.find_elements(By.CSS_SELECTOR, "a[href^='tel:']")
                            phone = ", ".join(sorted(set([p.text.strip() for p in phones if p.text.strip()])))
                        except:
                            pass

                        site = ""
                        try:
                            sites = driver.find_elements(By.CSS_SELECTOR, "a[target='_blank']")
                            for s in sites:
                                h = s.get_attribute("href")
                                if not h:
                                    continue
                                lower = h.lower()
                                if any(bad in lower for bad in ["2gis.ru", "vk.com", "ok.ru", "instagram.com", "t.me"]):
                                    continue
                                site = h
                                break
                        except:
                            pass

                        email = ""
                        source = ""
                        try:
                            mails = driver.find_elements(By.CSS_SELECTOR, "a[href^='mailto:']")
                            if mails:
                                email = mails[0].get_attribute("href").replace("mailto:", "").strip()
                                source = "2ГИС"
                        except:
                            pass

                        if not email:
                            if not site:
                                site = search_site_duckduckgo(name, f"{area} {REGION_NAME}")
                            if site:
                                email = find_emails_on_site_recursive(site)
                                if email:
                                    source = "Сайт"
                            if not email:
                                email = search_email_yandex_selenium(driver, f"{name} {area} {REGION_NAME}")
                                if email:
                                    source = "Яндекс"

                        # фильтр дублей
                        if email:
                            key = ("email", email.lower())
                        else:
                            key = ("name_phone", name.lower(), phone)

                        if key in seen_keys:
                            print(f"     SKIP (дубль) {name}")
                            continue
                        seen_keys.add(key)

                        print(f"     {name} | 📍 {area} | 📧 {email or '-'}")
                        ws.append([area, category, name, phone, site, email, source, link])
                        wb.save(OUTPUT_FILE)
                    except:
                        pass

                try:
                    next_url = f"{BASE_URL}/search/{quote(search_query)}/page/{page_num + 1}"
                    driver.get(next_url)
                    time.sleep(3)
                    check_captcha_smart(driver)
                    if f"/page/{page_num + 1}" not in driver.current_url:
                        break
                except:
                    break

    driver.quit()
    print(f"✅ Готово! Файл: {OUTPUT_FILE}")

if __name__ == "__main__":
    main()
